# > Imports
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, colors
from openpyxl.utils import get_column_letter

try:
    from openpyxl.cell import get_column_letter
except ImportError:
    from openpyxl.utils import get_column_letter
    from openpyxl.utils import column_index_from_string

# > Junker Drop
def junk_drop(SourceDF):
    trash = []
    SourceDF.dropna(subset=['MV Owner'], inplace=True)
    SourceDF = SourceDF.loc[SourceDF['Pay Status'] == 'On Hold']
    trash.append(SourceDF.loc[SourceDF['MV Owner'] == '-'])
    trash.append(SourceDF.loc[SourceDF['MV Owner'] == 'No current user'])
    for i in range(len(trash)):
        SourceDF = SourceDF.drop(trash[i].index)
    return SourceDF

# > Username data crossover
def user_cross(usernamesOnHold, usernamesSource):
    usernames = []
    for x in range(len(usernamesOnHold)):
        if (usernamesOnHold[x] in usernamesSource):
            usernames.append(usernamesOnHold[x])
    return usernames

# > Header format:
def format_header(path):
    wb = openpyxl.load_workbook(path)
    ws = wb['Sheet1']
    ws.title = 'Latam KPI AUTOGENERATOR(1.0)'
    fill_cell = PatternFill(patternType='solid', fgColor='C0DF16')

    for column_cells in ws.columns:
        new_column_length = max(len(str(cell.value)) for cell in column_cells)
        new_column_letter = (get_column_letter(column_cells[0].column))
        if new_column_length > 0:
            ws.column_dimensions[new_column_letter].width = new_column_length*1.23

    # > Cells letter order
    cell = []
    for i in range(ord('A'), ord('L')+1):
        cell.append((chr(i))+"1")

    for x in cell:
        ws[x].fill = fill_cell
        ws[x].font = Font(bold=True, color='ffffff', size='12')
    
    wb.save(path)
    wb.close()